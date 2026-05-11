#!/usr/bin/env python3
"""
seed_2026_from_excel.py

Reads ENERO–MAYO.xlsx from historical-data/ and generates
api/src/main/resources/db/migration/V12__seed_2026_operational_data.sql

Usage:
  cd lavadero-api
  python3 scripts/seed_2026_from_excel.py
"""

import datetime
import math
import re
from pathlib import Path

import openpyxl

# ── config ──────────────────────────────────────────────────────────────────
SCRIPT_DIR = Path(__file__).parent
REPO_ROOT = SCRIPT_DIR.parent
HISTORICAL_DIR = REPO_ROOT.parent / "historical-data"
OUTPUT_SQL = REPO_ROOT / "api/src/main/resources/db/migration/V12__seed_2026_operational_data.sql"

MONTH_MAP = {
    "ENERO": 1, "FEBRERO": 2, "MARZO": 3, "ABRIL": 4, "MAYO": 5,
    "JUNIO": 6, "JULIO": 7, "AGOSTO": 8, "SEPTIEMBRE": 9,
    "OCTUBRE": 10, "NOVIEMBRE": 11, "DICIEMBRE": 12,
}

FILES = [
    ("ENERO.xlsx",     1),
    ("FEBRERO.xlsx",   2),
    ("MARZO (2).xlsx", 3),
    ("ABRIL.xlsx",     4),
    ("MAYO.xlsx",      5),
]

SKIP_SHEETS = {"CORTE DOLARES", "HOJA MUESTRA"}
LAVADOR_DELIMITERS = re.compile(r"[+/&]|\bY\b")
GASTOS_ROW_IDX = 3   # 0-indexed (row 4 in Excel)
GASTOS_COL_IDX = 15  # 0-indexed (column P in Excel)


# ── helpers ──────────────────────────────────────────────────────────────────
def esc(s: str) -> str:
    return s.replace("'", "''")


def split_lavadores(raw: str) -> list[str]:
    parts = LAVADOR_DELIMITERS.split(raw)
    return [p.strip().upper() for p in parts if p.strip()]


def share_pcts(n: int) -> list[int]:
    if n == 1:
        return [100]
    base = 100 // n
    shares = [base] * n
    shares[0] += 100 - sum(shares)
    return shares


def is_ticket_row(r: list) -> bool:
    return (
        isinstance(r[0], int)
        and isinstance(r[2], int)
        and r[6] is not None
        and isinstance(r[6], (int, float))
    )


def parse_date(sheet_name: str, year: int = 2026) -> datetime.date | None:
    """Parse '01 ENERO' → date(2026, 1, 1)."""
    parts = sheet_name.strip().split()
    if len(parts) != 2:
        return None
    day_str, month_name = parts
    month = MONTH_MAP.get(month_name.upper())
    if month is None:
        return None
    try:
        return datetime.date(year, month, int(day_str))
    except ValueError:
        return None


def shift_type(t: datetime.time | None) -> str:
    if t is None or t < datetime.time(14, 0):
        return "MATUTINO"
    return "VESPERTINO"


# ── data extraction ───────────────────────────────────────────────────────────
def extract_all() -> tuple[list[datetime.date], dict, set[str]]:
    """Return (sorted_dates, days_data, all_employee_names).

    days_data[date] = {
        'tickets': [{'daily_seq', 'nota_number', 'vehicle', 'currency',
                     'amount', 'shift', 'courtesy', 'lavadores'}],
        'gastos': float | None,
    }
    """
    days_data: dict = {}
    all_employees: set[str] = set()
    used_notas: set[str] = set()  # global uniqueness for nota_number

    for filename, expected_month in FILES:
        path = HISTORICAL_DIR / filename
        if not path.exists():
            print(f"  WARNING: {path} not found — skipping")
            continue
        wb = openpyxl.load_workbook(path, data_only=True)
        for sheet_name in wb.sheetnames:
            if sheet_name in SKIP_SHEETS:
                continue
            d = parse_date(sheet_name)
            if d is None:
                print(f"  WARNING: could not parse date from '{sheet_name}' in {filename}")
                continue

            ws = wb[sheet_name]
            all_rows = list(ws.iter_rows(values_only=True))

            # Gastos del dia
            gastos = None
            try:
                g = all_rows[GASTOS_ROW_IDX][GASTOS_COL_IDX]
                if g is not None and isinstance(g, (int, float)) and float(g) > 0:
                    gastos = float(g)
            except IndexError:
                pass

            # Ticket rows start at Excel row 8 (index 7)
            tickets = []
            for row in all_rows[7:]:
                r = list(row)
                if not is_ticket_row(r):
                    continue
                weekly_seq = r[0]
                nota_raw = r[1]
                daily_seq = r[2]
                vehicle = str(r[3]).strip().upper() if r[3] else ""
                moneda = str(r[5]).strip().upper() if r[5] else ""
                importe = float(r[6])
                hora: datetime.time | None = r[7] if isinstance(r[7], datetime.time) else None
                lav_raw = str(r[8]).strip() if r[8] else ""

                courtesy = moneda not in ("MX", "USD") or importe == 0
                currency = "USD" if moneda == "USD" else "MXN"

                if nota_raw is not None and str(nota_raw).strip():
                    base_nota = str(nota_raw).strip()
                else:
                    base_nota = f"CORTESIA-{d.isoformat()}-{daily_seq}"
                nota_number = base_nota
                suffix = 2
                while nota_number in used_notas:
                    nota_number = f"{base_nota}-{suffix}"
                    suffix += 1
                used_notas.add(nota_number)

                lavadores = split_lavadores(lav_raw) if lav_raw else []
                all_employees.update(lavadores)

                tickets.append({
                    "daily_seq": daily_seq,
                    "nota_number": nota_number,
                    "vehicle": vehicle,
                    "currency": currency,
                    "amount": importe,
                    "shift": shift_type(hora),
                    "courtesy": courtesy,
                    "lavadores": lavadores,
                })

            days_data[d] = {"tickets": tickets, "gastos": gastos}

    sorted_dates = sorted(days_data.keys())
    return sorted_dates, days_data, all_employees


# ── SQL generation ─────────────────────────────────────────────────────────────
def generate_sql(sorted_dates: list, days_data: dict, all_employees: set[str]) -> str:
    lines: list[str] = []

    lines.append("-- V12: Seed 2026 operational data from Hojas Diarias (ENERO–MAYO)")
    lines.append(f"-- Generated by scripts/seed_2026_from_excel.py")
    lines.append(f"-- Days: {len(sorted_dates)} | Employees: {len(all_employees)}")
    lines.append("")

    # ── Block 1: Catalog ────────────────────────────────────────────────────
    lines.append("-- ═══════════════════════════════════════════════════")
    lines.append("-- Block 1: Catalog (HISTORICO service type + vehicle size)")
    lines.append("-- ═══════════════════════════════════════════════════")
    lines.append("")
    lines.append("INSERT INTO service_types (tenant_id, code, name, active)")
    lines.append("VALUES (1, 'HISTORICO', 'Historico', true)")
    lines.append("ON CONFLICT (tenant_id, code) DO NOTHING;")
    lines.append("")
    lines.append("INSERT INTO vehicle_sizes (tenant_id, code, name, sort_order, active)")
    lines.append("VALUES (1, 'HISTORICO', 'Historico', 99, true)")
    lines.append("ON CONFLICT (tenant_id, code) DO NOTHING;")
    lines.append("")
    lines.append("INSERT INTO service_prices (tenant_id, service_type_id, vehicle_size_id, amount, currency, effective_from)")
    lines.append("SELECT 1, st.id, vs.id, 1.00, 'MXN', '2026-01-01'")
    lines.append("FROM service_types st")
    lines.append("JOIN vehicle_sizes vs ON vs.code = 'HISTORICO' AND vs.tenant_id = 1")
    lines.append("WHERE st.code = 'HISTORICO' AND st.tenant_id = 1")
    lines.append("ON CONFLICT (tenant_id, service_type_id, vehicle_size_id, currency, effective_from) DO NOTHING;")
    lines.append("")

    # ── Block 2: Employees ──────────────────────────────────────────────────
    lines.append("-- ═══════════════════════════════════════════════════")
    lines.append("-- Block 2: Employees")
    lines.append("-- ═══════════════════════════════════════════════════")
    lines.append("")
    sorted_employees = sorted(all_employees)
    for name in sorted_employees:
        n = esc(name)
        lines.append(f"INSERT INTO employees (tenant_id, full_name, active)")
        lines.append(f"SELECT 1, '{n}', true")
        lines.append(f"WHERE NOT EXISTS (SELECT 1 FROM employees WHERE tenant_id = 1 AND full_name = '{n}');")
    lines.append("")

    # ── Block 3+: Per-day data ──────────────────────────────────────────────
    lines.append("-- ═══════════════════════════════════════════════════")
    lines.append("-- Block 3: Business days, shifts, tickets, assignments, expenses")
    lines.append("-- ═══════════════════════════════════════════════════")
    lines.append("")

    for d in sorted_dates:
        day = days_data[d]
        ds = d.isoformat()
        tickets = day["tickets"]
        gastos = day["gastos"]

        lines.append(f"-- ── {ds} ──────────────────────────")
        lines.append("")

        # Business day
        lines.append(f"INSERT INTO business_days (tenant_id, business_date, status)")
        lines.append(f"VALUES (1, '{ds}', 'CLOSED')")
        lines.append(f"ON CONFLICT (tenant_id, business_date) DO NOTHING;")
        lines.append("")

        # Shifts
        for shift in ("MATUTINO", "VESPERTINO"):
            lines.append(f"INSERT INTO shifts (tenant_id, business_day_id, shift_type, status)")
            lines.append(f"SELECT 1, id, '{shift}', 'CLOSED' FROM business_days")
            lines.append(f"WHERE business_date = '{ds}' AND tenant_id = 1")
            lines.append(f"ON CONFLICT (tenant_id, business_day_id, shift_type) DO NOTHING;")
            lines.append("")

        # Tickets
        for t in tickets:
            shift = t["shift"]
            daily_seq = t["daily_seq"]
            nota = esc(t["nota_number"])
            vehicle = esc(t["vehicle"])
            amount = t["amount"]
            currency = t["currency"]
            courtesy = "true" if t["courtesy"] else "false"
            courtesy_reason = ", 'Importado desde Excel'" if t["courtesy"] else ", NULL"

            lines.append(f"INSERT INTO tickets (")
            lines.append(f"  tenant_id, business_day_id, shift_id, service_type_id, vehicle_size_id,")
            lines.append(f"  daily_seq, nota_number, vehicle_description,")
            lines.append(f"  price_amount, currency, payment_method, courtesy, courtesy_reason, status")
            lines.append(f")")
            lines.append(f"SELECT")
            lines.append(f"  1, bd.id, s.id, st.id, vs.id,")
            lines.append(f"  {daily_seq}, '{nota}', '{vehicle}',")
            lines.append(f"  {amount:.2f}, '{currency}', 'CASH', {courtesy}{courtesy_reason}, 'ACTIVE'")
            lines.append(f"FROM business_days bd")
            lines.append(f"JOIN shifts s ON s.business_day_id = bd.id AND s.shift_type = '{shift}'")
            lines.append(f"JOIN service_types st ON st.code = 'HISTORICO' AND st.tenant_id = 1")
            lines.append(f"JOIN vehicle_sizes vs ON vs.code = 'HISTORICO' AND vs.tenant_id = 1")
            lines.append(f"WHERE bd.business_date = '{ds}' AND bd.tenant_id = 1")
            lines.append(f"ON CONFLICT (tenant_id, business_day_id, daily_seq) DO NOTHING;")
            lines.append("")

        # Ticket assignments
        for t in tickets:
            if not t["lavadores"]:
                continue
            lavs = t["lavadores"]
            shares = share_pcts(len(lavs))
            daily_seq = t["daily_seq"]
            ds_str = ds

            for lav, pct in zip(lavs, shares):
                lines.append(f"INSERT INTO ticket_assignments (tenant_id, ticket_id, employee_id, share_pct)")
                lines.append(f"SELECT 1, t.id, e.id, {pct}")
                lines.append(f"FROM tickets t")
                lines.append(f"JOIN business_days bd ON t.business_day_id = bd.id")
                lines.append(f"JOIN employees e ON e.full_name = '{esc(lav)}' AND e.tenant_id = 1")
                lines.append(f"WHERE bd.business_date = '{ds_str}' AND t.daily_seq = {daily_seq} AND bd.tenant_id = 1")
                lines.append(f"ON CONFLICT (tenant_id, ticket_id, employee_id) DO NOTHING;")
                lines.append("")

        # Expenses
        if gastos is not None and gastos > 0:
            lines.append(f"INSERT INTO expenses (tenant_id, business_day_id, expense_date, category, amount, description)")
            lines.append(f"SELECT 1, bd.id, '{ds}', 'OTHER', {gastos:.2f}, 'Gastos del dia (Excel import)'")
            lines.append(f"FROM business_days bd")
            lines.append(f"WHERE bd.business_date = '{ds}' AND bd.tenant_id = 1")
            lines.append(f"  AND NOT EXISTS (")
            lines.append(f"    SELECT 1 FROM expenses e")
            lines.append(f"    WHERE e.business_day_id = bd.id AND e.tenant_id = 1 AND e.description = 'Gastos del dia (Excel import)'")
            lines.append(f"  );")
            lines.append("")

    return "\n".join(lines)


# ── main ──────────────────────────────────────────────────────────────────────
def main():
    print("==> Extracting data from Excel files...")
    sorted_dates, days_data, all_employees = extract_all()

    total_tickets = sum(len(v["tickets"]) for v in days_data.values())
    print(f"    Days processed   : {len(sorted_dates)}")
    print(f"    Total tickets    : {total_tickets}")
    print(f"    Unique employees : {len(all_employees)}")
    print(f"    Date range       : {sorted_dates[0]} → {sorted_dates[-1]}")
    print()

    print("==> Generating SQL...")
    sql = generate_sql(sorted_dates, days_data, all_employees)

    OUTPUT_SQL.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_SQL.write_text(sql, encoding="utf-8")
    print(f"==> Written to {OUTPUT_SQL}")
    print(f"    File size: {OUTPUT_SQL.stat().st_size:,} bytes")
    print()
    print("Next steps:")
    print("  1. Review first 100 lines of V12 SQL")
    print("  2. docker compose up postgres -d")
    print("  3. cd api && ./mvnw spring-boot:run  (Flyway applies V12)")


if __name__ == "__main__":
    main()

#!/usr/bin/env python3
"""Load MARZO.xlsx into a local Lavadero API for dashboard testing.

This is intentionally a dev utility, not a Flyway migration. It creates catalog
rows and tickets through the public API so local data exercises the real app.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import urllib.error
import urllib.request
from dataclasses import dataclass
from datetime import date
from decimal import Decimal
from pathlib import Path

from openpyxl import load_workbook


DEFAULT_XLSX = "/Users/brandonrubio/Desktop/Lavado AI Script/MARZO.xlsx"


@dataclass(frozen=True)
class Row:
    business_date: date
    nota: str
    auto: str
    currency: str
    amount: Decimal
    lavadores: tuple[str, ...]


class ApiClient:
    def __init__(self, base_url: str) -> None:
        self.base_url = base_url.rstrip("/")

    def get(self, path: str):
        return self._request("GET", path)

    def post(self, path: str, payload: dict):
        return self._request("POST", path, payload)

    def _request(self, method: str, path: str, payload: dict | None = None):
        data = None if payload is None else json.dumps(payload).encode("utf-8")
        request = urllib.request.Request(
            f"{self.base_url}{path}",
            data=data,
            method=method,
            headers={"Content-Type": "application/json"},
        )
        try:
            with urllib.request.urlopen(request, timeout=20) as response:
                body = response.read()
                if not body:
                    return None
                return json.loads(body.decode("utf-8"))
        except urllib.error.HTTPError as exc:
            body = exc.read().decode("utf-8")
            try:
                parsed = json.loads(body)
                message = parsed.get("error", body)
            except json.JSONDecodeError:
                message = body
            raise RuntimeError(f"{method} {path} failed: {exc.code} {message}") from exc


def parse_rows(path: Path, year: int, limit: int | None) -> list[Row]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    rows: list[Row] = []
    for sheet in workbook.worksheets:
        match = re.fullmatch(r"(\d{2}) MARZO", sheet.title.strip().upper())
        if not match:
            continue
        business_date = date(year, 3, int(match.group(1)))
        for cells in sheet.iter_rows(min_row=8, values_only=True):
            nota, auto, currency, amount, lavadores = cells[1], cells[3], cells[5], cells[6], cells[8]
            if not nota or not auto or not currency or amount in (None, "") or not lavadores:
                continue
            parsed_lavadores = split_lavadores(str(lavadores))
            if not parsed_lavadores:
                continue
            rows.append(Row(
                business_date=business_date,
                nota=str(nota),
                auto=str(auto).strip(),
                currency=normalize_currency(str(currency)),
                amount=Decimal(str(amount)).quantize(Decimal("0.01")),
                lavadores=parsed_lavadores,
            ))
            if limit and len(rows) >= limit:
                return rows
    return rows


def split_lavadores(value: str) -> tuple[str, ...]:
    cleaned = value.upper().replace("&", ",").replace("/", ",").replace("+", ",")
    cleaned = re.sub(r"\s+Y\s+", ",", cleaned)
    names = [name.strip(" .-") for name in cleaned.split(",")]
    return tuple(dict.fromkeys(name for name in names if name))


def normalize_currency(value: str) -> str:
    upper = value.strip().upper()
    if upper in {"MX", "MEX", "MEXICANO", "MXN"}:
        return "MXN"
    if upper in {"US", "USD", "DOLAR", "DOLARES"}:
        return "USD"
    raise ValueError(f"Unsupported currency: {value}")


def slug(value: str) -> str:
    normalized = re.sub(r"[^A-Z0-9]+", "_", value.upper()).strip("_")
    return normalized[:34] or "ITEM"


def first_by(items: list[dict], key: str, value):
    return next((item for item in items if item.get(key) == value), None)


def ensure_employee(api: ApiClient, employees: dict[str, int], name: str) -> int:
    if name in employees:
        return employees[name]
    created = api.post("/api/v1/employees", {"fullName": name.title()})
    employees[name] = created["id"]
    return created["id"]


def ensure_service_type(api: ApiClient) -> int:
    code = "MARZO_IMPORT"
    services = api.get("/api/v1/service-types")
    existing = first_by(services, "code", code)
    if existing:
        return existing["id"]
    created = api.post("/api/v1/service-types", {
        "code": code,
        "name": "Lavado Marzo importado",
        "description": "Servicio usado solo para cargar MARZO.xlsx en desarrollo",
    })
    return created["id"]


def ensure_vehicle_size(api: ApiClient, sizes_by_code: dict[str, int], currency: str, amount: Decimal) -> int:
    cents = int(amount * 100)
    code = f"IMP_{currency}_{cents}"
    if code in sizes_by_code:
        return sizes_by_code[code]
    created = api.post("/api/v1/vehicle-sizes", {
        "code": code,
        "name": f"Importe {currency} {amount}",
        "sortOrder": cents,
    })
    sizes_by_code[code] = created["id"]
    return created["id"]


def ensure_price(api: ApiClient, prices_seen: set[tuple[int, int, str, Decimal]], service_type_id: int,
                 vehicle_size_id: int, currency: str, amount: Decimal, effective_from: date) -> None:
    key = (service_type_id, vehicle_size_id, currency, amount)
    if key in prices_seen:
        return
    try:
        api.post("/api/v1/service-prices", {
            "serviceTypeId": service_type_id,
            "vehicleSizeId": vehicle_size_id,
            "amount": str(amount),
            "currency": currency,
            "effectiveFrom": effective_from.isoformat(),
        })
    except RuntimeError as exc:
        if "Request conflicts with existing data" not in str(exc):
            raise
    prices_seen.add(key)


def ensure_business_day(api: ApiClient, business_date: date) -> int:
    path = f"/api/v1/business-days?from={business_date.isoformat()}&to={business_date.isoformat()}"
    existing = api.get(path)
    if existing:
        return existing[0]["id"]
    return api.post("/api/v1/business-days/open", {"businessDate": business_date.isoformat()})["id"]


def ensure_shift(api: ApiClient, business_day_id: int) -> int:
    existing = api.get(f"/api/v1/shifts?business_day_id={business_day_id}")
    matutino = first_by(existing, "shiftType", "MATUTINO")
    if matutino:
        return matutino["id"]
    return api.post("/api/v1/shifts/open", {"businessDayId": business_day_id, "shiftType": "MATUTINO"})["id"]


def main() -> int:
    parser = argparse.ArgumentParser(description="Load MARZO.xlsx into the local Lavadero API.")
    parser.add_argument("--file", default=DEFAULT_XLSX)
    parser.add_argument("--base-url", default="http://localhost:8080")
    parser.add_argument("--year", type=int, default=2026)
    parser.add_argument("--limit", type=int, default=75, help="Limit tickets loaded; use 0 for all rows.")
    args = parser.parse_args()

    path = Path(args.file)
    if not path.exists():
        print(f"File not found: {path}", file=sys.stderr)
        return 1

    rows = parse_rows(path, args.year, None if args.limit == 0 else args.limit)
    if not rows:
        print("No ticket rows found.", file=sys.stderr)
        return 1

    api = ApiClient(args.base_url)
    employees = {item["fullName"].upper(): item["id"] for item in api.get("/api/v1/employees?active=true")}
    sizes_by_code = {item["code"]: item["id"] for item in api.get("/api/v1/vehicle-sizes")}
    service_type_id = ensure_service_type(api)
    prices_seen: set[tuple[int, int, str, Decimal]] = set()

    created_tickets = 0
    for row in rows:
        employee_ids = [ensure_employee(api, employees, name) for name in row.lavadores]
        vehicle_size_id = ensure_vehicle_size(api, sizes_by_code, row.currency, row.amount)
        ensure_price(api, prices_seen, service_type_id, vehicle_size_id, row.currency, row.amount, row.business_date)
        business_day_id = ensure_business_day(api, row.business_date)
        shift_id = ensure_shift(api, business_day_id)
        api.post("/api/v1/tickets", {
            "businessDayId": business_day_id,
            "shiftId": shift_id,
            "serviceTypeId": service_type_id,
            "vehicleSizeId": vehicle_size_id,
            "currency": row.currency,
            "vehicleDescription": row.auto,
            "courtesy": False,
            "employeeIds": employee_ids,
        })
        created_tickets += 1

    dates = sorted({row.business_date.isoformat() for row in rows})
    print(f"Loaded {created_tickets} tickets from {path.name}.")
    print(f"Dates: {dates[0]} through {dates[-1]}")
    print(f"Open the dashboard for {dates[0]} to verify real imported activity.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
